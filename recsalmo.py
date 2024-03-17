# Copyrights 2023: Nuttachat Wisittipanit

import os
import itertools
import util
import time
from openpyxl import (Workbook, load_workbook, )
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import pandas as pd
from Bio import Phylo
from Bio import SeqIO
from Bio import Align
from Bio.SeqRecord import SeqRecord
from Bio.Phylo.TreeConstruction import (DistanceMatrix, DistanceTreeConstructor, )
from const import (File, ResultKey, Record, )
from structobj import (RES_ALL, RecordRES, AssemblyAnnotation, STAnnotation, SeroAnnotation, AssemblyAnnXcel)
from main import WGSSal
from binfutil import PerformListAlignment

import warnings

warnings.filterwarnings("ignore")


class BatchRun:

    def __init__(self, input_path, output_path, mainName=None):
        # self.useInputFolder = File.inputFolder
        # self.useOutputFolder = File.outputFolder
        self.useInputFolder = input_path
        self.useOutputFolder = output_path
        self.mainName = mainName
        self.fileCls = None  # File ins
        self.wgsFileList = []
        self.numGenomes = 0
        self.resultHash = {}  # {'BK_SAL1': RES_ALL ins,'BK_SAL4': RES_ALL ins, ..}
        self.genomeNameList = []  # list of genome names run in order
        self.headLabels = Record.headLabels
        self.recordTitle = Record.recordTitle
        self.recordRESList = []  # RecordRES ins
        self.spProfileList = []  # list of SpacerProfile ins

    def setupFolderPath(self):
        """
        Setup all folder paths for the program
        :return:
        """
        self.fileCls = File(self.useInputFolder, self.useOutputFolder, self.mainName)
        self.fileCls.setupAllPaths()

    def databaseInitialize(self):
        """
        Initialize some databases
        :return:
        """
        # copy 'sptemp.fasta' from 'spacerall.fasta' if not already exists
        util.checkCopySpacerFasta(self.fileCls)
        self.makeSpacerDB()
        self.makeSpacerTempDB()

    def collectInfo(self):
        """
        Collect information from all the input files
        :return:
        """
        self.wgsFileList = [f for f in os.listdir(self.useInputFolder) if
                            os.path.isfile(os.path.join(self.useInputFolder, f))]
        self.wgsFileList.sort()
        self.numGenomes = len(self.wgsFileList)

    def run(self):
        """
        Main run method
        :return:
        """
        # clear DBSC folder first
        util.createFolderAndClear(self.fileCls.DBSC_Folder)
        for i, wgsFile in enumerate(self.wgsFileList):
            wSal = WGSSal(self.fileCls, wgsFile)
            wSal.setupIO()
            # create folder if not already existing
            util.createFolderNotExist(wSal.outFolderPath)
            # if wSal.outFolderName not in ('SAL_YC8093AA_AS','BK_SAL1', ):
            #     continue
            self.makeBlastDB(wSal.inputFilePath, wSal.outFolderName)
            wSal.runMLST()
            wSal.runSISTR()
            wSal.runAMRFinderPlus()
            wSal.runSPIFinder()
            wSal.runCrisprFinder()
            self.genomeNameList.append((wSal.outFolderName, wSal.outFolderPath))
            self.collectResult(wSal)
            if i == 3:
                break
        # Perform Spacer Profile Alignment
        finalMatrix = self.performSpacerProfileAlignment()
        if len(finalMatrix) != 0:
            self.constructPhylogeneticTreeSpacerProfile(finalMatrix)
        self.removeTempSpacerDB()
        # Record all results in the excel file
        self.recordResult()
        # Read from the record
        self.readFromRecord()
        # Construct pie charts
        self.constructPieChartEach('serovar')
        # self.constructPieChartEach('identification')
        self.constructPieChartEach('call_ST')
        # Construct ParSNP tree
        self.constructSNPTree()
        # Remove all folders inside /DBSC/
        util.createFolderAndClear(self.fileCls.DBSC_Folder)

    def run_CRISPR(self):
        """
        Main run method
        :return:
        """
        # clear DBSC folder first
        util.createFolderAndClear(self.fileCls.DBSC_Folder)
        for i, wgsFile in enumerate(self.wgsFileList):
            wSal = WGSSal(self.fileCls, wgsFile)
            wSal.setupIO()
            # Test only
            # if wSal.outFolderName not in ('SAL_BA3995AA_AS', ):
            #     continue
            # if wSal.outFolderName not in ('SAL_HC6463AA_AS', ):
            #     continue
            # if wSal.outFolderName not in ('SAL_EA2914AA_AS', ):
            #     continue
            # if wSal.outFolderName not in ('SAL_KA9071AA_AS', ):
            #     continue
            # if wSal.outFolderName not in ('SAL_HC4765AA_AS', 'SAL_HC4766AA_AS', 'SAL_FA1054AA_AS', ):
            #     continue
            # if wSal.outFolderName not in ('SAL_VC9662AA_AS', ):
            #     continue
            # if wSal.outFolderName not in ('SAL_KA9071AA_AS', 'SAL_LA0087AA_AS','SAL_QC3063AA_AS'):
            #     continue
            # if wSal.outFolderName not in ('SAL_BA3995AA_AS', 'SAL_BA
            # 804AA_AS', 'SAL_HC6481AA_AS','SAL_IB8216AA_AS'):
            #     continue
            # if wSal.outFolderName not in (
            # 'SAL_BA3995AA_AS', 'SAL_BA5804AA_AS', 'SAL_HC6463AA_AS', 'SAL_HC6481AA_AS', 'SAL_IB8216AA_AS',
            # 'SAL_EA2914AA_AS','SAL_KA9071AA_AS', 'SAL_LA0087AA_AS','SAL_QC3063AA_AS'):
            #     continue
            # if wSal.outFolderName not in ('SAL_BA3995AA_AS', 'SAL_BA5804AA_AS', 'SAL_EA2914AA_AS'):
            #     continue
            # if wSal.outFolderName not in ('SAL_BA3995AA_AS','SAL_BA5804AA_AS','SAL_HC6463AA_AS','SAL_HC6481AA_AS'):
            #     continue
            # make folder right here
            util.createFolderAndClear(wSal.outFolderPath)
            # make DB in the "/maindb/DBSC/BK_SAL1_DB"
            self.makeBlastDB(wSal.inputFilePath, wSal.outFolderName)
            # call runCrisprFinder method
            wSal.runCrisprFinder()
            self.genomeNameList.append((wSal.outFolderName, wSal.outFolderPath))
            self.collectResult_CRISPR(wSal)
            # if i == 50:
            #     break
        # # Perform Spacer Profile Alignment
        # finalMatrix = self.performSpacerProfileAlignment()
        # if len(finalMatrix) != 0:
        #     self.constructPhylogeneticTreeSpacerProfile(finalMatrix)
        self.removeTempSpacerDB()
        # Record all results in the excel file
        self.recordResult_CRISPR()
        # Remove all folders inside /DBSC/
        util.createFolderAndClear(self.fileCls.DBSC_Folder)

    def aggregateST(self):
        """
        We need to aggregate
        :return:
        """
        main_dir = '/home/nuttachat/Downloads/SalmoStrain'
        xcelFilePath = os.path.join(main_dir, 'summary_THAICR.xlsx')
        assemFilePath = os.path.join(main_dir, 'Thai_1.txt')
        stFilePath = os.path.join(main_dir, 'ThaiST_1.txt')
        seroFilePath = os.path.join(main_dir, 'Thai_SERO.txt')
        assemList = util.read_file_tab(assemFilePath, headInclude=True)
        stList = util.read_file_tab(stFilePath, headInclude=True)
        seroList = util.read_file_tab(seroFilePath, headInclude=True)
        # self.assemHash -> {Uberstrain: AssemblyAnnotation ins,..}
        self.assemHash = {}
        # self.assemHashX -> {Assembly Barcode: AssemblyAnnotation ins,..}
        self.assemHashX = {}
        # self.stHash -> {Uberstrain: STAnnotation ins,..}
        self.stHash = {}
        # self.seroHash -> {Uberstrain: STAnnotation ins,..}
        self.seroHash = {}
        for k in assemList:
            # k -> ['SAL_AA0066AA','184756',..,'SAL_AA0066AA_AS'] ->
            #      [Uberstrain=0,Name=1,Serovar=18,..,Assembly Barcode=38]
            uberStrain = k[0]
            serovar = k[19]
            assemBarcode = k[40]
            assemAnn = AssemblyAnnotation(uberStrain, assemBarcode, serovar)
            self.assemHash[uberStrain] = assemAnn
            self.assemHashX[assemBarcode] = assemAnn
        for k in stList:
            # k -> ['SAL_AA0066AA','184756',..,'SAL_AA0066AA_AS'] ->
            #      [Uberstrain=0,..,ST=34]
            uberStrain = k[0]
            stIndex = k[33]
            stAnn = STAnnotation(uberStrain, stIndex)
            self.stHash[uberStrain] = stAnn
        for k in seroList:
            # k -> ['SAL_AA0066AA','184756',..,'SAL_AA0066AA_AS'] ->
            #      [Uberstrain=0,..,Sero='Typhi']
            uberStrain = k[0]
            serotype = k[32]
            seroAnn = SeroAnnotation(uberStrain, serotype)
            self.seroHash[uberStrain] = seroAnn
        print('assemHash')
        print(self.assemHash)
        for key in self.assemHash:
            print(self.assemHash[key].AssemblyBarcode)
            print(self.assemHash[key].Serovar)
            break
        print('assemHashX')
        print(self.assemHashX)
        # for key in self.assemHashX:
        #     print(self.assemHashX[key].AssemblyBarcode)
        #     print(self.assemHashX[key].Serovar)
        print('stHash')
        print(self.stHash)
        for key in self.stHash:
            print(self.stHash[key].STIndex)
            break
        # Read xcelFilePath
        wb = load_workbook(xcelFilePath)
        ws = wb['summary']
        rows = list(ws.rows)
        headerOld = []
        dataOld = []
        crpLen = []  # [(C1 len,C2 len),..]
        for i in range(len(rows)):
            b = []
            row = rows[i]
            if i == 0:
                for col in range(len(row)):
                    value = row[col].value
                    headerOld.append(value)
            else:
                for col in range(len(row)):
                    value = row[col].value
                    b.append(value)
                dataOld.append(b)
                # row[5] -> Pos-C1, row[6] -> Pos-C2
                posC1 = str(row[5].value)  # 3077831-3078162
                posC2 = str(row[6].value)  # 3077831-3078162
                # find length
                if posC1 != 'None-None':
                    # print(posC1.split('-'))
                    v1 = posC1.split('-')
                    v1 = list(filter(lambda x: x != '', v1))
                    startC1 = int(v1[0])
                    stopC1 = int(v1[1])
                    lenC1 = abs(startC1 - stopC1) + 1
                else:
                    lenC1 = 'None'
                if posC2 != 'None-None':
                    v2 = posC2.split('-')
                    v2 = list(filter(lambda x: x != '', v2))
                    startC2 = int(v2[0])
                    stopC2 = int(v2[1])
                    lenC2 = abs(startC2 - stopC2) + 1
                else:
                    lenC2 = 'None'
                crpLen.append((lenC1, lenC2))
        # print('headerOld')
        # print(headerOld)
        # print('dataOld')
        # print(dataOld[0])
        headerNew = headerOld + ['Serovar', 'STIndex', 'Serotype', 'LenC1(bp)', 'LenC2(bp)']
        dataNew = []
        for i, a in enumerate(dataOld):
            # a -> ['a','b','c','d',..]
            # barcode -> 'SAL_AA0066AA_AS'
            barcode = a[0]
            # gg = []
            serovar = self.assemHashX[barcode].Serovar
            uber = self.assemHashX[barcode].uberStrain
            st = self.stHash[uber].STIndex
            serotype = self.seroHash[uber].serotype
            lenC1 = crpLen[i][0]
            lenC2 = crpLen[i][1]
            gg = a + [serovar, st, serotype, lenC1, lenC2]
            dataNew.append(gg)
        print(dataNew[0])
        # CREATE NEW WORKBOOK
        wb = Workbook()
        ws_sum = wb.active
        ws_sum.title = 'main'
        ws_sum.column_dimensions[get_column_letter(1)].width = 15
        ws_sum.column_dimensions[get_column_letter(2)].width = 30
        ws_sum.column_dimensions[get_column_letter(3)].width = 25
        currentRow = 1
        for i, label in enumerate(headerNew):
            ws_sum.cell(row=currentRow, column=i + 1, value=label)
        # currentRow = 2
        for i, data in enumerate(dataNew):
            # data -> [a,b,c,d]
            for j, a in enumerate(data):
                ws_sum.cell(row=i + 2, column=j + 1, value=a)

        newFilePath = os.path.join(main_dir, 'summary_THAICR_NEWX.xlsx')
        wb.save(newFilePath)

    def crisprPrimerAlignment(self):
        """
        We need to aggregate
        :return:
        """
        main_dir = '/home/aonlazio/Python/Project/WGSCP'
        inputFolderPath = os.path.join(main_dir, 'input')
        outputFolderPath = os.path.join(main_dir, 'output')
        resFolderPath = os.path.join(main_dir, 'res')
        xceloutFolderPath = os.path.join(resFolderPath, 'xcelout')
        xcelFilePath = os.path.join(xceloutFolderPath, 'summary_THAICR_NEWX.xlsx')
        primerFilePath = os.path.join(resFolderPath, 'primer.fasta')
        # read primer
        for seqRecord in SeqIO.parse(primerFilePath, "fasta"):
            if seqRecord.id == 'forward_C1':
                primerForward_C1 = seqRecord
            elif seqRecord.id == 'reverse_C1':
                primerReverse_C1 = seqRecord
            elif seqRecord.id == 'forward_C2':
                primerForward_C2 = seqRecord
            elif seqRecord.id == 'reverse_C2':
                primerReverse_C2 = seqRecord

        # Read xcelFilePath
        wb = load_workbook(xcelFilePath)
        ws = wb['main']
        rows = list(ws.rows)
        headerOld = []
        dataOld = []
        for i in range(len(rows)):
            b = []
            row = rows[i]
            if i == 0:
                for col in range(len(row)):
                    value = row[col].value
                    headerOld.append(value)
            else:
                for col in range(len(row)):
                    value = row[col].value
                    b.append(value)
                dataOld.append(b)

        # assemHash -> {AssemblyName: AssemblyAnnotation ins,..}
        assemHash = {}
        for i, a in enumerate(dataOld):
            # a -> ['a','b','c','d',..]
            # assemblyName -> 'SAL_AA0066AA_AS'
            name = a[0]
            spacerC1 = a[1]
            spacerC2 = a[2]
            nodeC1 = a[9]
            nodeC2 = a[10]
            assemHash[name] = AssemblyAnnXcel(name)
            assemHash[name].spacerC1 = spacerC1
            assemHash[name].spacerC2 = spacerC2
            assemHash[name].nodeC1 = nodeC1
            assemHash[name].nodeC2 = nodeC2

        # print(assemHash)
        # for a in assemHash:
        #     print(assemHash[a].AssemblyName)
        #     print(assemHash[a].spacerC1)
        #     print(assemHash[a].spacerC2)
        #     print(assemHash[a].nodeC1)
        #     print(assemHash[a].nodeC2)

        aligner = Align.PairwiseAligner()
        # aligner.substitution_matrix = substitution_matrices.load("BLOSUM62")
        aligner.mode = 'local'
        aligner.match_score = 2
        aligner.mismatch_score = -1
        aligner.open_gap_score = -3
        aligner.extend_gap_score = -0.5

        # iterate files over input folder
        wgsFileList = [f for f in os.listdir(inputFolderPath) if
                       os.path.isfile(os.path.join(inputFolderPath, f))]

        # get node in FASTA in SeqRecord ins
        # print(wgsFileList)
        for wgsFile in wgsFileList:
            # wgsFile -> SAL_BA4239AA_AS.scaffold.fasta
            # print(wgsFile)
            assemName = wgsFile.split('.')[0]
            wgsFilePath = os.path.join(inputFolderPath, wgsFile)
            # print('assemName: ', assemName)
            # print('wgsFilePath: ', wgsFilePath)
            # print('nodeC1: ', assemHash[assemName].nodeC1)
            # print('nodeC2: ', assemHash[assemName].nodeC2)
            nodeC1rec = None  # SeqRecord ins
            nodeC2rec = None  # SeqRecord ins
            try:
                # get SeqRecord from nodeC1
                for seqRecord in SeqIO.parse(wgsFilePath, "fasta"):
                    if seqRecord.id == assemHash[assemName].nodeC1:
                        nodeC1rec = seqRecord
                        break
                # get SeqRecord from nodeC2
                for seqRecord in SeqIO.parse(wgsFilePath, "fasta"):
                    if seqRecord.id == assemHash[assemName].nodeC2:
                        nodeC2rec = seqRecord
                        break
            except:
                continue
            # print('primerForward_C1: ', primerForward_C1.seq)
            # print('primerReverse_C1: ', primerReverse_C1.seq)
            # print('primerForward_C2: ', primerForward_C2.seq)
            # print('primerReverse_C2: ', primerReverse_C2.seq)
            # print('nodeC1rec: ', nodeC1rec.seq)
            # print('nodeC2rec: ', nodeC2rec.seq)
            # print('len(nodeC1rec.seq): ', len(nodeC1rec.seq))
            # print('len(nodeC2rec.seq): ', len(nodeC2rec.seq))
            try:
                nodeC1LEN = len(nodeC1rec.seq)
                nodeC2LEN = len(nodeC2rec.seq)
            except:
                continue

            A_CRISPR1_L, A_CRISPR2_L = self.determineCrisprLength(nodeC1rec.seq,
                                                                  nodeC1rec.seq.reverse_complement(),
                                                                  primerForward_C1.seq,
                                                                  primerReverse_C1.seq,
                                                                  nodeC2rec.seq,
                                                                  nodeC2rec.seq.reverse_complement(),
                                                                  primerForward_C2.seq,
                                                                  primerReverse_C2.seq,
                                                                  aligner,
                                                                  assemHash[assemName])
            B_CRISPR1_L, B_CRISPR2_L = self.determineCrisprLength(nodeC1rec.seq.reverse_complement(),
                                                                  nodeC1rec.seq,
                                                                  primerForward_C1.seq,
                                                                  primerReverse_C1.seq,
                                                                  nodeC2rec.seq.reverse_complement(),
                                                                  nodeC2rec.seq,
                                                                  primerForward_C2.seq,
                                                                  primerReverse_C2.seq,
                                                                  aligner,
                                                                  assemHash[assemName])
            print('xxx')
            print('A_CRISPR1_L: ', A_CRISPR1_L)
            # print('A_CRISPR2_L: ', A_CRISPR2_L)
            # print('B_CRISPR1_L: ', B_CRISPR1_L)
            # print('B_CRISPR2_L: ', B_CRISPR2_L)
            assemHash[assemName].A_CRISPR1_L = A_CRISPR1_L
            assemHash[assemName].A_CRISPR2_L = A_CRISPR2_L
            assemHash[assemName].B_CRISPR1_L = B_CRISPR1_L
            assemHash[assemName].B_CRISPR2_L = B_CRISPR2_L
            # break

        headerNew = headerOld + ['LPRIMER_C1_A', 'LPRIMER_C2_A', 'LPRIMER_C1_B', 'LPRIMER_C2_B', ]
        dataNew = []
        for i, a in enumerate(dataOld):
            # a -> ['a','b','c','d',..]
            # assemName -> 'SAL_AA0066AA_AS'
            assemName = a[0]
            A_CRISPR1_L = assemHash[assemName].A_CRISPR1_L
            A_CRISPR2_L = assemHash[assemName].A_CRISPR2_L
            B_CRISPR1_L = assemHash[assemName].B_CRISPR1_L
            B_CRISPR2_L = assemHash[assemName].B_CRISPR2_L
            gg = a + [A_CRISPR1_L, A_CRISPR2_L, B_CRISPR1_L, B_CRISPR2_L]
            dataNew.append(gg)
        # CREATE NEW WORKBOOK
        wb = Workbook()
        ws_sum = wb.active
        ws_sum.title = 'summary'
        ws_sum.column_dimensions[get_column_letter(1)].width = 15
        ws_sum.column_dimensions[get_column_letter(2)].width = 30
        ws_sum.column_dimensions[get_column_letter(3)].width = 25
        currentRow = 1
        for i, label in enumerate(headerNew):
            ws_sum.cell(row=currentRow, column=i + 1, value=label)
        # currentRow = 2
        for i, data in enumerate(dataNew):
            # data -> [a,b,c,d]
            for j, a in enumerate(data):
                ws_sum.cell(row=i + 2, column=j + 1, value=a)

        newFilePath = os.path.join(outputFolderPath, 'summary_PRIMER2.xlsx')
        wb.save(newFilePath)

    def generateFileFolderFromSerotype(self):
        """
        Group all Assembly files using Serotype
        (1) For each folder, make FASTA file
          - SAL_AB1316AA_AS_NODEC1.fasta : contain only NODE sequence
          - SAL_AB1316AA_AS_NODEC2.fasta : contain only NODE sequence
        :return:
        """
        main_dir = '/home/aonlazio/Python/Project/WGSCP'
        inputFolderPath = os.path.join(main_dir, 'input')
        outputFolderPath = os.path.join(main_dir, 'output')
        resFolderPath = os.path.join(main_dir, 'res')
        # xceloutFolderPath = os.path.join(resFolderPath, 'xcelout')
        xcelreadFilePath = os.path.join(outputFolderPath, 'summary_PRIMER2.xlsx')
        assemoutFolderPath = os.path.join(resFolderPath, 'assemout')
        util.createFolderAndClear(assemoutFolderPath)

        AssemblyList = []
        SerotypeList = []

        # Read xcelFilePath
        wb = load_workbook(xcelreadFilePath)
        ws = wb['summary']
        rows = list(ws.rows)
        headerOld = []
        dataOld = []
        for i in range(len(rows)):
            b = []
            row = rows[i]
            if i == 0:
                for col in range(len(row)):
                    value = row[col].value
                    headerOld.append(value)
            else:
                for col in range(len(row)):
                    value = row[col].value
                    b.append(value)
                dataOld.append(b)
        # assemHash -> {AssemblyName: AssemblyAnnotation ins,..}
        assemHash = {}
        for i, a in enumerate(dataOld):
            # a -> ['a','b','c','d',..]
            # assemblyName -> 'SAL_AA0066AA_AS'
            name = a[0]
            # spacerC1 = a[1]
            # spacerC2 = a[2]
            nodeC1 = a[9]
            nodeC2 = a[10]
            Serovar = a[11]
            Serotype = a[13]
            assemHash[name] = AssemblyAnnXcel(name)
            assemHash[name].nodeC1 = nodeC1
            assemHash[name].nodeC2 = nodeC2
            assemHash[name].Serotype = Serotype
            assemHash[name].Serovar = Serovar
            SerotypeList.append(Serotype)
        # SerotypeGroupList -> ['Newport','Stanley',..]
        SerotypeGroupList = list(set(SerotypeList))
        SerotypeGroupList.sort()  # sort letters
        print('SerotypeGroupList: ', SerotypeGroupList)

        # iterate files over input folder
        wgsFileList = [f for f in os.listdir(inputFolderPath) if
                       os.path.isfile(os.path.join(inputFolderPath, f))]
        # get node in FASTA in SeqRecord ins
        # print(wgsFileList)
        for i, wgsFile in enumerate(wgsFileList):
            # wgsFile -> SAL_BA4239AA_AS.scaffold.fasta
            # print(wgsFile)
            # if i == 4:
            #     break
            assemName = wgsFile.split('.')[0]
            try:
                Serotype = assemHash[assemName].Serotype
            except (KeyError,):
                print('error assemName: ', assemName)
                continue
            # create folder for Serotype
            # serotypeFolderPath -> folder name 'Stanley'
            serotypeFolderPath = os.path.join(assemoutFolderPath, Serotype)
            # util.createFolderAndClear(serotypeFolderPath)
            try:
                os.mkdir(serotypeFolderPath)
            except (FileExistsError,):
                # folder already exists
                pass

            wgsFilePath = os.path.join(inputFolderPath, wgsFile)
            nodeC1rec = None  # SeqRecord ins
            nodeC2rec = None  # SeqRecord ins
            try:
                # get SeqRecord from nodeC1
                for seqRecord in SeqIO.parse(wgsFilePath, "fasta"):
                    if seqRecord.id == assemHash[assemName].nodeC1:
                        nodeC1rec = seqRecord
                        break
                # get SeqRecord from nodeC2
                for seqRecord in SeqIO.parse(wgsFilePath, "fasta"):
                    if seqRecord.id == assemHash[assemName].nodeC2:
                        nodeC2rec = seqRecord
                        break
            except:
                continue
            if nodeC1rec is not None and nodeC2rec is not None:
                fastaC1FileName = '%s_NODEC1.fasta' % (assemName,)
                fastaC2FileName = '%s_NODEC2.fasta' % (assemName,)
                fastaC1FilePath = os.path.join(serotypeFolderPath, fastaC1FileName)
                fastaC2FilePath = os.path.join(serotypeFolderPath, fastaC2FileName)
                with open(fastaC1FilePath, "w") as output_handle:
                    SeqIO.write([nodeC1rec], output_handle, "fasta")
                with open(fastaC2FilePath, "w") as output_handle:
                    SeqIO.write([nodeC2rec], output_handle, "fasta")

    def determineAlignPositions(self, element):
        """
        element -> [[5461 5468]
                    [5469 5487]]
        :param element:
        :return:
        """
        start = None  # 5461
        stop = None  # 5487
        for i, a in enumerate(element):
            # i -> 0 or 1
            # a -> [5461 5468] or [5469 5487]
            if i == 0:
                start = a[0]
            if i == len(element) - 1:
                # len(RVC2[0]) -> 2
                stop = a[1]
        return start, stop

    def determineCrisprLength(self,
                              nodeFWC1seq,
                              nodeRVC1seq,
                              primerFWC1seq,
                              primerRVC1seq,
                              nodeFWC2seq,
                              nodeRVC2seq,
                              primerFWC2seq,
                              primerRVC2seq,
                              aligner,
                              AAX):
        """
        determineCrisprLength
        :param nodeFWC1seq: node C1 forward sequence
        :param nodeRVC1seq: node C1 reverse sequence
        :param primerFWC1seq: primer C1 forward sequence
        :param primerRVC1seq: primer C1 reverse sequence
        :param nodeFWC2seq: node C2 forward sequence
        :param nodeRVC2seq: node C2 reverse sequence
        :param primerFWC2seq: primer C2 forward sequence
        :param primerRVC2seq: primer C2 reverse sequence
        :param aligner: aligner
        :param AAX: AssemblyAnnXcel ins
        :return: C1_L and C2_L
        """
        nodeC1LEN = len(nodeFWC1seq)
        nodeC2LEN = len(nodeFWC2seq)
        alignFWC1s = aligner.align(nodeFWC1seq, primerFWC1seq)
        alignRVC1s = aligner.align(nodeRVC1seq, primerRVC1seq)
        alignFWC2s = aligner.align(nodeFWC2seq, primerFWC2seq)
        alignRVC2s = aligner.align(nodeRVC2seq, primerRVC2seq)
        alignFWC1 = alignFWC1s[0]
        alignRVC1 = alignRVC1s[0]
        alignFWC2 = alignFWC2s[0]
        alignRVC2 = alignRVC2s[0]
        FWC1 = alignFWC1.aligned
        RVC1 = alignRVC1.aligned
        FWC2 = alignFWC2.aligned
        RVC2 = alignRVC2.aligned
        # print(RVC2)
        # print(type(FWC1))
        # RVC2[0] -> [[82999 83016]
        #             [83017 83025]]
        # RVC2[1] -> [[    0    17]
        #             [   17    25]]
        # Extract Positions RVC2
        start_FWC1, stop_FWC1 = self.determineAlignPositions(FWC1[0])
        start_RVC1, stop_RVC1 = self.determineAlignPositions(RVC1[0])
        start_FWC2, stop_FWC2 = self.determineAlignPositions(FWC2[0])
        start_RVC2, stop_RVC2 = self.determineAlignPositions(RVC2[0])
        # print('start_FWC1: ', start_FWC1)
        # print('stop_FWC1: ', stop_FWC1)
        # print('start_RVC1: ', start_RVC1)
        # print('stop_RVC1: ', stop_RVC1)
        # print('start_FWC2: ', start_FWC2)
        # print('stop_FWC2: ', stop_FWC2)
        # print('start_RVC2: ', start_RVC2)
        # print('stop_RVC2: ', stop_RVC2)
        # Adjust start_RVC1 and stop_RVC1
        start_RVC1_adj = nodeC1LEN - stop_RVC1
        stop_RVC1_adj = nodeC1LEN - start_RVC1
        # print('start_RVC1_adj: ', start_RVC1_adj)
        # print('stop_RVC1_adj: ', stop_RVC1_adj)
        start_RVC2_adj = nodeC2LEN - stop_RVC2
        stop_RVC2_adj = nodeC2LEN - start_RVC2
        # print('start_RVC2_adj: ', start_RVC2_adj)
        # print('stop_RVC2_adj: ', stop_RVC2_adj)
        CRISPR1_L = stop_RVC1_adj - start_FWC1
        CRISPR2_L = stop_RVC2_adj - start_FWC2
        # AAX.CRISPR1_L = CRISPR1_L
        # AAX.CRISPR2_L = CRISPR2_L
        # print('CRISPR1_L: ', AAX.CRISPR1_L)
        # print('CRISPR2_L: ', AAX.CRISPR2_L)
        return CRISPR1_L, CRISPR2_L

    def run_AMRandSPI(self):
        """
        Main run method
        :return:
        """
        # clear DBSC folder first
        util.createFolderAndClear(self.fileCls.DBSC_Folder)
        for i, wgsFile in enumerate(self.wgsFileList):
            wSal = WGSSal(self.fileCls, wgsFile)
            wSal.setupIO()
            # Test only
            # if wSal.outFolderName not in ('SAL_BA3995AA_AS', 'SAL_BA5799AA_AS',):
            #     continue
            # make folder right here
            util.createFolderAndClear(wSal.outFolderPath)
            # make DB in the "/maindb/DBSC/BK_SAL1_DB"
            self.makeBlastDB(wSal.inputFilePath, wSal.outFolderName)
            # call runCrisprFinder method
            wSal.runAMRFinderPlus()
            wSal.runSPIFinder()
            self.genomeNameList.append((wSal.outFolderName, wSal.outFolderPath))
            self.collectResult_AMRandSPI(wSal)
        # Record all results in the excel file
        self.recordResult_AMRandSPI()
        # Remove all folders inside /DBSC/
        util.createFolderAndClear(self.fileCls.DBSC_Folder)

    def constructSNPTree(self):
        """
        Construct SNP-based tree
        :return:
        """
        cmdLine = 'parsnp -c -r %s -d %s -o %s' % (self.fileCls.refGenome_LT2_FilePath,
                                                   self.useInputFolder,
                                                   self.fileCls.treeOutputFolder,
                                                   )
        os.system(cmdLine)

    def collectResult(self, wSal):
        """
        Collect all results
        :return:
        """
        self.resultHash[wSal.outFolderName] = RES_ALL(wSal.outFolderName)
        # res -> RES_ALL ins
        resAll = self.resultHash[wSal.outFolderName]
        # resAll.res_SeqSero2 -> RES_SeqSero2 ins
        # resAll.res_SeqSero2 = wSal.seqsero2_res
        resAll.res_MLST = wSal.mlst_res
        resAll.res_SISTR = wSal.sistr_res
        resAll.res_AMRFinder = wSal.amrfinder_res
        resAll.res_SPI = wSal.spifinder_res
        resAll.res_CRISPR = wSal.crisprfinder_res
        # after all results are collected, delete the temp folder
        tempDir_SISTR = os.path.join(self.fileCls.outputFolder, '%s_%s' % (wSal.outFolderName, 'tmpSistr',))
        util.removeFolderAndContent(tempDir_SISTR)

    def collectResult_CRISPR(self, wSal):
        """
        Collect all results
        :return:
        """
        # self.resultHash -> {'BK_SAL1: RES_ALL ins}
        self.resultHash[wSal.outFolderName] = RES_ALL(wSal.outFolderName)
        # res -> RES_ALL ins
        resAll = self.resultHash[wSal.outFolderName]
        # print('resAll: ', resAll)
        resAll.res_CRISPR = wSal.crisprfinder_res
        # print('resAll.res_CRISPR: ', resAll.res_CRISPR)

    def collectResult_AMRandSPI(self, wSal):
        """
        Collect all results
        :return:
        """
        # self.resultHash -> {'BK_SAL1: RES_ALL ins}
        self.resultHash[wSal.outFolderName] = RES_ALL(wSal.outFolderName)
        # res -> RES_ALL ins
        resAll = self.resultHash[wSal.outFolderName]
        print('resAll: ', resAll)
        resAll.res_AMRFinder = wSal.amrfinder_res
        resAll.res_SPI = wSal.spifinder_res
        print('resAll.res_AMRFinder: ', resAll.res_AMRFinder)
        print('resAll.res_SPI: ', resAll.res_SPI)

    def makeBlastDB(self, scaffoldFilePath, genomeName):
        """
        Need to transform a scaffolds file into database for the search to begin
        For instance, 'BK_SAL1' is associated with the file 'BK_SAL1.scaffolds.fasta'
        This one file is split into 8 sub-files which are all put into "/maindb/DBSC/BK_SAL1" folder.
        :param scaffoldFilePath: e.g. 'in_file/BK_SAL1.scaffolds.fasta'
        :param genomeName: e.g. 'BK_SAL1' or 'SAL_IC4008AA_AS'
        :return:
        """
        srList = []
        for seq_record in SeqIO.parse(scaffoldFilePath, "fasta"):
            try:
                newID = seq_record.id[0:50]
            except:
                newID = seq_record.id
            record = SeqRecord(seq_record.seq, id=newID, description='')
            srList.append(record)
        # Write into the file
        # fastaFilePath = os.path.join(scaffoldFilePath, 'NodeFound.fasta')
        with open(scaffoldFilePath, "w") as output_handle:
            SeqIO.write(srList, output_handle, "fasta")
        # create main db folder e.g. "/maindb/DBSC/BK_SAL1_DB"
        genomeName_DB = '%s_DB' % (genomeName,)  # 'BK_SAL1_DB'
        # dbFolder -> "/maindb/DBSC/BK_SAL1_DB"
        dbFolder = os.path.join(self.fileCls.DBSC_Folder, genomeName_DB)
        util.createFolderAndClear(dbFolder)
        # dbFolder -> "/maindb/DBSC/BK_SAL1_DB/BK_SAL1"
        dbFolder_X = os.path.join(dbFolder, genomeName)
        cmdLine = 'makeblastdb -in %s -dbtype nucl -parse_seqids -out %s' % (scaffoldFilePath,
                                                                             dbFolder_X,
                                                                             )
        os.system(cmdLine)

    def makeRefLT2_DB(self):
        """
        Need to transform a scaffolds file into database for the search to begin
        For instance, 'BK_SAL1' is associated with the file 'BK_SAL1.scaffolds.fasta'
        This one file is split into 8 sub-files which are all put into "/maindb/DBSC/BK_SAL1" folder.
        :param scaffoldFilePath: e.g. 'in_file/BK_SAL1.scaffolds.fasta'
        :param genomeName: e.g. 'BK_SAL1' or 'SAL_IC4008AA_AS'
        :return:
        """
        RefGenome_DB = 'LT2DB'
        # dbFolder -> "/maindb/REF/LT2_DB"
        dbFolder = os.path.join(self.fileCls.REFDBFolder, RefGenome_DB)
        util.createFolderAndClear(dbFolder)
        # dbFolder -> "/maindb/REF/LT2_DB/BK_SAL1"
        dbFolder_X = os.path.join(dbFolder, 'LT2')
        cmdLine = 'makeblastdb -in %s -dbtype nucl -parse_seqids -out %s' % (self.fileCls.refGenome_LT2_FilePath,
                                                                             dbFolder_X,
                                                                             )
        os.system(cmdLine)

    def makeSpacerDB(self):
        """
        Transform a 'spacerall.fasta' into the database used in Blast operation
        :param genomeName: genome name e.g. 'BK_SAL1'
        :return:
        """
        # get the list of scaffolds files in "in_file" folder
        spacerFile = 'spacerall.fasta'  # Spacer file
        spacerFilePath = os.path.join(self.fileCls.SPACER_Folder, spacerFile)
        # create main db folder -> "/SPACER/SP_DB"
        spacerDB_name = 'SP_DB'
        # dbFolder -> "/SPACER/SP_DB"
        dbFolder = os.path.join(self.fileCls.SPACER_Folder, spacerDB_name)
        if os.path.exists(dbFolder) is False:
            # DB not exist yet, create new one
            util.createFolderAndClear(dbFolder)
            # dbFolder -> "/SPACER/SP_DB/SPALL"
            dbFolder_X = os.path.join(dbFolder, 'SPALL')
            cmdLine = 'makeblastdb -in %s -dbtype nucl -parse_seqids -out %s' % (spacerFilePath,
                                                                                 dbFolder_X,
                                                                                 )
            os.system(cmdLine)

    def makeSpacerTempDB(self):
        """
        Transform a 'spacerall.fasta' into the database used in Blast operation
        :param genomeName: genome name e.g. 'BK_SAL1'
        :return:
        """
        # get the list of scaffolds files in "in_file" folder
        spacerFile = 'spacerall.fasta'  # Spacer file
        spacerFilePath = os.path.join(self.fileCls.SPACER_Folder, spacerFile)
        # create main db folder -> "/SPACER/SP_DB"
        spacerDB_name = 'SPTEMP_DB'
        # dbFolder -> "/SPACER/SPTEMP_DB"
        dbFolder = os.path.join(self.fileCls.SPACER_Folder, spacerDB_name)
        util.createFolderAndClear(dbFolder)
        # dbFolder -> "/SPACER/SP_DB/SPALL"
        dbFolder_X = os.path.join(dbFolder, 'SPALL')
        cmdLine = 'makeblastdb -in %s -dbtype nucl -parse_seqids -out %s' % (spacerFilePath,
                                                                             dbFolder_X,
                                                                             )
        os.system(cmdLine)

    def performSpacerProfileAlignment(self):
        """
        Perform spacer profile alignment among all genomes
        :return:
        """
        # print(self.genomeNameList)
        for genomeName, folderPath in self.genomeNameList:
            spProfileFilePath = os.path.join(folderPath, 'spProfile.txt')
            try:
                spProfile = util.readSpacerProfile(spProfileFilePath)
            except (FileNotFoundError,):
                continue
            self.spProfileList.append((genomeName, spProfile))
        # find all combinations in pairs
        matrix = []
        finalMatrix = []
        # comb -> [(a,b),(c,d),..]
        combList = list(itertools.combinations(self.spProfileList, 2))
        if len(combList) == 0:
            return finalMatrix
        n = 0
        j = 0
        k = 0
        startNumber = len(self.genomeNameList)  # if len = 5, n -> 4
        u = []
        while True:
            # a -> ('BK_SAL1',spProfile)
            # b -> ('BK_SAL10',spProfile)
            a, b = combList[n]
            # align C1 first
            pair_C1 = PerformListAlignment(a[1].C1, b[1].C1)
            pair_C1.run()
            pair_C2 = PerformListAlignment(a[1].C2, b[1].C2)
            pair_C2.run()
            totalScore = pair_C1.totalScore + pair_C2.totalScore
            u.append(totalScore)
            n += 1
            j += 1
            if j == startNumber - 1:
                startNumber -= 1
                j = 0
                u = [0] * (k + 1) + u
                k += 1
                matrix.append(u)
                u = []
            if n == len(combList):
                matrix.append([0] * (len(self.genomeNameList)))
                break
        # transform to numpy array
        X = np.array(matrix)
        # copy to lower triangle
        X = np.triu(X)
        X = X + X.T - np.diag(np.diag(X))
        # get only the lower triangle
        for i in range(X.shape[0]):
            ab = X[i][0:i + 1]
            ab = ab.tolist()
            finalMatrix.append(list(ab))
        return finalMatrix

    def constructPhylogeneticTreeSpacerProfile(self, distanceMatrix):
        """
        Construct the tree by UPGMA method
        :param distanceMatrix:
        :return:
        """

        def get_label(leaf):
            a = str(leaf.name)
            if a.find('Inner') != -1:
                # find 'Inner'
                return None
            return leaf.name

        pureList = [u[0] for u in self.genomeNameList]
        # print(pureList)
        b = DistanceMatrix(names=pureList, matrix=distanceMatrix)
        # Creating a DistanceTreeConstructor object
        constructor = DistanceTreeConstructor()
        # Construct the phlyogenetic tree using UPGMA algorithm
        UPGMATree = constructor.upgma(b)
        # Draw the phlyogenetic tree
        # Phylo.draw(UPGMATree,axes=fig_axes, do_show=False)
        plt.clf()
        # fig1, ax1 = plt.subplots(figsize=(11, 11))
        fig = plt.figure(figsize=(15, 20), dpi=150)
        fig_axes = fig.add_subplot(1, 1, 1)
        Phylo.draw(UPGMATree, label_func=get_label, axes=fig_axes, do_show=False)
        # save tree to PNG file
        plt.savefig(self.fileCls.CRISPR_TreeFilePath)

    def recordResult(self):
        """
        Record result to excel/csv
        :return:
        """
        print('recordResult')
        wb = Workbook()
        ws_sum = wb.active
        ws_sum.title = self.recordTitle
        ws_sum.column_dimensions[get_column_letter(1)].width = 15
        ws_sum.column_dimensions[get_column_letter(2)].width = 30
        ws_sum.column_dimensions[get_column_letter(3)].width = 25
        currentRow = 1
        for i, label in enumerate(self.headLabels):
            ws_sum.cell(row=currentRow, column=i + 1, value=label)
        currentRow = 2
        for name in self.resultHash:
            resAll = self.resultHash[name]  # RES_ALL ins
            rc = resAll.res_CRISPR
            spC1 = rc.spacerC1
            spC2 = rc.spacerC2
            num_spC1 = len(spC1.split('|')) if spC1 != '' else 0
            num_spC2 = len(spC2.split('|')) if spC2 != '' else 0
            print('rc.posEndC1: ', type(rc.posEndC1))
            print('rc.posStartC1: ', rc.posStartC1)
            print('rc.posEndC2: ', rc.posEndC2)
            print('rc.posStartC2: ', rc.posStartC2)
            try:
                lenC1 = rc.posEndC1 - rc.posStartC1
            except (TypeError,):
                lenC1 = "NA"
            try:
                lenC2 = rc.posEndC2 - rc.posStartC2
            except (TypeError,):
                lenC2 = "NA"
            resList = [name,
                       resAll.res_SISTR.cgmlst_subspecies,
                       resAll.res_SISTR.serovar,
                       resAll.res_MLST.call_ST,  # MLST
                       resAll.res_SISTR.cgmlst_ST,  # cgMLST
                       resAll.res_AMRFinder.amrClass,  # AMR gene/classes/subclasses
                       resAll.res_SPI.spiList,  # SPI info
                       spC1,  # spacers of Crispr locus 1
                       spC2,  # spacers of Crispr locus 2
                       num_spC1,
                       num_spC2,
                       rc.DR_OneC1,
                       rc.DR_OneC2,
                       "%s-%s" % (rc.posStartC1, rc.posEndC1,),
                       "%s-%s" % (rc.posStartC2, rc.posEndC2,),
                       lenC1,
                       lenC2,
                       ]
            for j, resValue in enumerate(resList):
                ws_sum.cell(row=currentRow, column=j + 1, value=resValue)
            currentRow += 1
        wb.save(self.fileCls.recordFilePath)
        # convert to CSV
        r = pd.read_excel(self.fileCls.recordFilePath, sheet_name=self.recordTitle)
        r.to_csv(self.fileCls.recordFilePathCSV, index=None, header=True)

    def recordResult_CRISPR(self):
        """
        Record result to excel/csv
        :return:
        """
        wb = Workbook()
        # summary sheet
        ws_sum = wb.active
        ws_sum.title = self.recordTitle
        ws_sum.column_dimensions[get_column_letter(1)].width = 15
        ws_sum.column_dimensions[get_column_letter(2)].width = 30
        ws_sum.column_dimensions[get_column_letter(3)].width = 25
        currentRow = 1
        for i, label in enumerate(Record.headLabels_CRISPR):
            ws_sum.cell(row=currentRow, column=i + 1, value=label)
        currentRow = 2
        for name in self.resultHash:
            resAll = self.resultHash[name]  # RES_ALL ins
            resList = [name,
                       resAll.res_CRISPR.spacerC1,  # spacers of Crispr locus 1
                       resAll.res_CRISPR.spacerC2,  # spacers of Crispr locus 2
                       ]
            for j, resValue in enumerate(resList):
                ws_sum.cell(row=currentRow, column=j + 1, value=resValue)
            currentCol = len(resList) + 1
            spC1 = resAll.res_CRISPR.spacerC1
            spC2 = resAll.res_CRISPR.spacerC2
            C1_len = len(spC1.split('|')) if spC1 != '' else 0
            C2_len = len(spC2.split('|')) if spC2 != '' else 0
            ws_sum.cell(row=currentRow, column=currentCol, value=C1_len)
            ws_sum.cell(row=currentRow, column=currentCol + 1, value=C2_len)
            rc = resAll.res_CRISPR
            ws_sum.cell(row=currentRow, column=currentCol + 2, value="%s-%s" % (rc.posStartC1, rc.posEndC1,))
            ws_sum.cell(row=currentRow, column=currentCol + 3, value="%s-%s" % (rc.posStartC2, rc.posEndC2,))
            ws_sum.cell(row=currentRow, column=currentCol + 4, value=rc.DR_OneC1)
            ws_sum.cell(row=currentRow, column=currentCol + 5, value=rc.DR_OneC2)
            ws_sum.cell(row=currentRow, column=currentCol + 6, value=rc.NodeC1)
            ws_sum.cell(row=currentRow, column=currentCol + 7, value=rc.NodeC2)
            currentRow += 1
        # spacer sheet
        currentRow = 1
        ws_sp = wb.create_sheet('spacer')
        for name in self.resultHash:
            resAll = self.resultHash[name]  # RES_ALL ins
            spC1 = resAll.res_CRISPR.spacerC1
            spC2 = resAll.res_CRISPR.spacerC2
            spC1_s = spC1.split('|')
            spC2_s = spC2.split('|')
            ws_sp.cell(row=currentRow, column=1, value=name)
            currentCol = 2
            for k, sp in enumerate(spC1_s):
                ws_sp.cell(row=currentRow, column=k + currentCol, value=sp)
            for k, sp in enumerate(spC2_s):
                ws_sp.cell(row=currentRow + 1, column=k + currentCol, value=sp)
            currentRow += 2
        wb.save(self.fileCls.recordFilePath)

    def recordResult_AMRandSPI(self):
        """
        Record result to excel/csv
        :return:
        """
        wb = Workbook()
        # summary sheet
        ws_sum = wb.active
        ws_sum.title = self.recordTitle
        ws_sum.column_dimensions[get_column_letter(1)].width = 15
        ws_sum.column_dimensions[get_column_letter(2)].width = 50
        ws_sum.column_dimensions[get_column_letter(3)].width = 50
        currentRow = 1
        for i, label in enumerate(Record.headLabels_AMRandSPI):
            ws_sum.cell(row=currentRow, column=i + 1, value=label)
        currentRow = 2
        for name in self.resultHash:
            resAll = self.resultHash[name]  # RES_ALL ins
            resList = [name,
                       resAll.res_AMRFinder.amrClass_Gene,  # tet(B)|...
                       resAll.res_SPI.spiList,  # SPI1|SPI2|SPI3
                       ]
            for j, resValue in enumerate(resList):
                ws_sum.cell(row=currentRow, column=j + 1, value=resValue)
            currentRow += 1
        wb.save(self.fileCls.recordFilePath_AMRandSPI)

    def readFromRecord(self):
        """
        Read summary result from record
        :return:
        """
        wb = load_workbook(self.fileCls.recordFilePath)
        ws = wb[self.recordTitle]
        rows = list(ws.rows)
        # data_x = []
        # data_y = []
        for i in range(len(rows)):
            if i == 0:
                # skip row 0
                continue
            row = rows[i]
            genomeName = row[0].value
            recordRES = RecordRES(genomeName)
            recordRES.subspecies = row[1].value
            recordRES.serovar = row[2].value
            recordRES.call_ST = row[3].value
            recordRES.cgmlst_ST = row[4].value
            # recordRES.serovar_cgmlst = row[5].value
            # recordRES.serogroup = row[6].value
            # recordRES.cgmlst_subspecies = row[7].value
            recordRES.amrClass = row[5].value
            self.recordRESList.append(recordRES)

    def constructPieChartEach(self, mainName):
        """
        Construct pie chart for each name
        :param mainName: 'serotype','identification'
        :return:
        """
        # create pie chart for serotype
        # define labels
        # count
        countDict = {}  # {'I 4,[5],12':1, 'Agona':2, 'Weltevreden':2}
        # labelSerotype = []  # ['I 4,[5],12','Agona','Weltevreden']
        for record in self.recordRESList:
            # record -> RecordRES ins
            # keyUse = record.serotype
            keyUse = getattr(record, mainName)
            if countDict.get(keyUse) is None:
                # no key yet, initialize with 1
                countDict[keyUse] = 1
            else:
                countDict[keyUse] += 1
        labelType = list(countDict.keys())
        lavelValues = list(countDict.values())
        sumA = sum(lavelValues)
        prList = []
        prSumSoFar = 0.0
        # calculate percentage of each serotype
        for i, name in enumerate(labelType):
            # name -> 'Agona'
            # get the value
            rawValue = countDict[name]  # 2
            if i != len(labelType):
                # NOT last item
                prValue = round(rawValue / sumA, 2)
                prSumSoFar += prValue
            else:
                # last item
                prValue = 100.0 - prSumSoFar
            prList.append(prValue)
        # define Seaborn color palette to use
        colors = sns.color_palette('pastel')[0:len(labelType)]
        plt.clf()
        fig1, ax1 = plt.subplots(figsize=(9, 9))
        # fig1.subplots_adjust(0.3, 0, 1, 1)
        plt.pie(prList, labels=labelType, colors=colors, autopct='%.0f%%', textprops={'fontsize': 11})
        # plt.show()
        saveFile = 'pieChart_%s.png' % (mainName,)
        # saveFile = 'pieChart_%s_test.png' % (mainName,)
        saveFilePath = os.path.join(self.fileCls.outputFolder, saveFile)
        plt.savefig(saveFilePath)

    def removeTempSpacerDB(self):
        """
        Delete the folder /SPACER/SPTEMP_DB and file /SPACER/sptemp.fasta
        :return:
        """
        dbFolder = os.path.join(self.fileCls.SPACER_Folder, 'SPTEMP_DB')
        util.removeFolderAndContent(dbFolder)
        spFastaTempFilePath = os.path.join(self.fileCls.SPACER_Folder, 'sptemp.fasta')
        util.removeFile(spFastaTempFilePath)


if __name__ == '__main__':
    import argparse
    from pathlib import Path

    parser = argparse.ArgumentParser(
        prog='RECSALMO',
        description='Specialized Python Program for Typing and Characterization of Salmonella Genomes',
        epilog='Thank you for using RECSALMO. Have a superb day!')
    requirement = parser.add_argument_group("requirement")
    requirement.add_argument("input_dir", help='Directory containing Salmonella genome assemblies')
    requirement.add_argument("output_dir", help='Main output directory')
    args = parser.parse_args()
    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)
    if not input_dir.exists():
        # if the input_dir does not exist, raise error
        print("The input directory does not exist")
        raise SystemExit(1)
    if not output_dir.exists():
        # if the output_dir does not exist, create one
        print("The output directory does not exist, creating output folder path %s .." % (output_dir,))
        print('output_dir: ', output_dir)
        os.mkdir(output_dir)
    start_time = time.time()
    obj = BatchRun(input_dir, output_dir)
    obj.setupFolderPath()
    obj.databaseInitialize()
    obj.collectInfo()
    obj.run()
    time_elapse = time.time() - start_time
    print("RECSALMO has finished the analysis of %s Salmonella genomes in %s seconds." % (obj.numGenomes, time_elapse))
    avgTime_genome = round(time_elapse / obj.numGenomes, 2)
    print("Average analysis time per genome equals %s seconds." % (avgTime_genome,))

    # # Test Only
    # # input_dir = '/home/nuttachat/Downloads/SalStrains/USA'
    # # output_dir = '/home/nuttachat/Downloads/SalStrains/USA_OUTV4'
    # # name = 'USA'
    # # input_dir = '/home/nuttachat/Downloads/SalmoStrain/THAI'
    # # output_dir = '/home/nuttachat/Downloads/SalmoStrain/THAI_OUT'
    # # name = 'THAICR'
    # input_dir = '/home/aonlazio/Python/Project/WGSCP38/input'
    # # output_dir = '/home/aonlazio/Python/Project/WGSCP38/output'
    # output_dir = '/home/aonlazio/Python/Project/WGSCP38/outputx1'
    # name = 'WGS40F'
    #
    # obj = BatchRun(input_dir, output_dir, mainName=name)
    # obj.setupFolderPath()
    # obj.databaseInitialize()
    # obj.collectInfo()
    # obj.run()
    # # obj.makeRefLT2_DB()
    # # obj.run_CRISPR()
    # # obj.aggregateST()
    # # obj.crisprPrimerAlignment()
    # # obj.generateFileFolderFromSerotype()
